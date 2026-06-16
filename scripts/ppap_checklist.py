#!/usr/bin/env python3
"""
PPAP检查清单生成工具
Usage: python3 ppap_checklist.py --level <1-5> --customer <OEM名称> --output <output_path>
"""

import argparse
import json
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# PPAP 18要素定义
PPAP_ELEMENTS = [
    {
        'id': 1,
        'name': '设计记录',
        'name_en': 'Design Records',
        'description': '工程图样、规范、CAD数据等设计文件',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '是否包含所有相关的设计文件？',
            '图样版本是否最新？',
            'CAD数据是否与图样一致？',
            '是否标识了特殊特性？'
        ]
    },
    {
        'id': 2,
        'name': '工程更改文件',
        'name_en': 'Engineering Change Documents',
        'description': '临时更改、ECN等',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '是否有未关闭的ECN？',
            '临时更改是否在有效期内？',
            '更改是否已落实到生产？'
        ]
    },
    {
        'id': 3,
        'name': '客户工程批准',
        'name_en': 'Customer Engineering Approval',
        'description': '如客户要求，需获得工程批准',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '是否需要客户工程批准？',
            '批准是否已获得？',
            '批准是否在有效期内？'
        ]
    },
    {
        'id': 4,
        'name': '设计FMEA',
        'name_en': 'Design FMEA',
        'description': '设计失效模式及后果分析',
        'levels': [3, 4, 5],
        'verification_points': [
            'DFMEA是否按AIAG-VDA七步法完成？',
            '高风险项(AP=H)是否有措施？',
            '特殊特性是否在DFMEA中识别？',
            'DFMEA是否由跨职能团队完成？'
        ]
    },
    {
        'id': 5,
        'name': '过程流程图',
        'name_en': 'Process Flow Diagram',
        'description': '制造过程流程图',
        'levels': [3, 4, 5],
        'verification_points': [
            '流程图是否覆盖所有制造步骤？',
            '是否包含检验/测试点？',
            '是否与PFMEA和控制计划一致？'
        ]
    },
    {
        'id': 6,
        'name': '过程FMEA',
        'name_en': 'Process FMEA',
        'description': '过程失效模式及后果分析',
        'levels': [3, 4, 5],
        'verification_points': [
            'PFMEA是否按AIAG-VDA七步法完成？',
            '所有过程步骤是否覆盖？',
            '高风险项(AP=H/M)是否有措施？',
            'PFMEA是否与控制计划关联？'
        ]
    },
    {
        'id': 7,
        'name': '控制计划',
        'name_en': 'Control Plan',
        'description': '试生产/生产控制计划',
        'levels': [3, 4, 5],
        'verification_points': [
            '控制计划是否包含所有过程步骤？',
            '特殊特性是否标识？',
            '控制方法是否与PFMEA一致？',
            '反应计划是否明确？',
            '是否包含初始过程研究要求？'
        ]
    },
    {
        'id': 8,
        'name': '测量系统分析研究',
        'name_en': 'Measurement System Analysis Studies',
        'description': 'MSA/Gage R&R研究',
        'levels': [3, 4, 5],
        'verification_points': [
            'MSA是否覆盖控制计划中所有测量系统？',
            'Gage R&R是否≤10%（或≤30%视应用）？',
            'ndc是否≥5？',
            '偏倚/线性/稳定性是否评估？'
        ]
    },
    {
        'id': 9,
        'name': '尺寸结果',
        'name_en': 'Dimensional Results',
        'description': '全尺寸检验结果',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '是否覆盖图样上所有尺寸？',
            '所有尺寸是否在公差范围内？',
            '特殊特性尺寸是否标注？',
            'GD&T要求是否检验？'
        ]
    },
    {
        'id': 10,
        'name': '材料/性能试验结果',
        'name_en': 'Material/Performance Test Results',
        'description': '材料和性能试验结果',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '材料试验是否满足规范要求？',
            '性能试验是否完成？',
            '试验是否由合格实验室完成？'
        ]
    },
    {
        'id': 11,
        'name': '初始过程研究',
        'name_en': 'Initial Process Study',
        'description': 'SPC/Cpk初始过程能力研究',
        'levels': [3, 4, 5],
        'verification_points': [
            'Cpk是否≥1.33（特殊特性≥1.67）？',
            '数据是否来自有效生产运行？',
            '过程是否稳定（控制图无异常）？',
            '样本量是否足够？'
        ]
    },
    {
        'id': 12,
        'name': '合格实验室文件',
        'name_en': 'Qualified Laboratory Documentation',
        'description': '实验室资质文件（ISO 17025等）',
        'levels': [3, 4, 5],
        'verification_points': [
            '是否有ISO 17025认证或等效？',
            '实验室范围是否覆盖所需试验？',
            '认证是否在有效期内？'
        ]
    },
    {
        'id': 13,
        'name': '外观批准报告(AAR)',
        'name_en': 'Appearance Approval Report',
        'description': '外观件批准报告（如适用）',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '产品是否有外观要求？',
            'AAR是否已获得客户批准？',
            '色板/样件是否保留？'
        ]
    },
    {
        'id': 14,
        'name': '样品生产件',
        'name_en': 'Sample Production Parts',
        'description': '代表性样品',
        'levels': [2, 3, 4, 5],
        'verification_points': [
            '样品是否来自有效生产运行？',
            '样品数量是否满足客户要求？',
            '样品是否正确标识？'
        ]
    },
    {
        'id': 15,
        'name': '主样品',
        'name_en': 'Master Samples',
        'description': '主样品及标识',
        'levels': [3, 4, 5],
        'verification_points': [
            '主样品是否保留？',
            '是否与生产件一致？',
            '标识是否正确？'
        ]
    },
    {
        'id': 16,
        'name': '检具和试验设备',
        'name_en': 'Checking Aids and Test Equipment',
        'description': '测量设备清单及校准记录',
        'levels': [3, 4, 5],
        'verification_points': [
            '检具是否通过MSA验证？',
            '校准是否在有效期内？',
            '检具是否覆盖所有特殊特性？'
        ]
    },
    {
        'id': 17,
        'name': '客户特殊要求',
        'name_en': 'Customer Specific Requirements',
        'description': 'CSR符合性记录',
        'levels': [3, 4, 5],
        'verification_points': [
            '是否获取并理解客户CSR？',
            '所有CSR是否满足？',
            'CSR文件是否可追溯？'
        ]
    },
    {
        'id': 18,
        'name': '零件提交保证书(PSW)',
        'name_en': 'Part Submission Warrant',
        'description': 'PPAP提交总表',
        'levels': [1, 2, 3, 4, 5],
        'verification_points': [
            'PSW信息是否完整？',
            '所有要素结果是否填写？',
            '声明是否签署？',
            '提交等级是否正确？'
        ]
    }
]


def generate_checklist_xlsx(level, customer, output_path):
    """生成PPAP检查清单Excel文件"""
    if not HAS_OPENPYXL:
        print("错误: openpyxl未安装，请运行: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"PPAP检查清单 Level {level}"
    
    # 样式定义
    title_font = Font(name='黑体', size=16, bold=True, color='003366')
    header_font = Font(name='黑体', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    sub_header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    normal_font = Font(name='宋体', size=10)
    required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f'PPAP检查清单 - Level {level}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # 信息行
    ws.merge_cells('A2:B2')
    ws['A2'] = '客户:'
    ws['C2'] = customer
    ws.merge_cells('D2:E2')
    ws['D2'] = '日期:'
    ws['F2'] = datetime.now().strftime('%Y-%m-%d')
    ws.merge_cells('G2:H2')
    ws['G2'] = '零件编号:'
    
    # 表头
    headers = ['序号', '要素名称', '英文名称', 'Level要求', '状态', '验证要点', '备注', '责任人']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 列宽
    widths = [6, 18, 25, 12, 10, 45, 20, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # 数据行
    row = 5
    for elem in PPAP_ELEMENTS:
        is_required = level in elem['levels']
        
        cell_id = ws.cell(row=row, column=1, value=elem['id'])
        cell_id.font = normal_font
        cell_id.alignment = center_align
        cell_id.border = thin_border
        
        cell_name = ws.cell(row=row, column=2, value=elem['name'])
        cell_name.font = Font(name='宋体', size=10, bold=is_required)
        cell_name.alignment = left_align
        cell_name.border = thin_border
        
        cell_name_en = ws.cell(row=row, column=3, value=elem['name_en'])
        cell_name_en.font = normal_font
        cell_name_en.alignment = left_align
        cell_name_en.border = thin_border
        
        cell_level = ws.cell(row=row, column=4, value=f'Level {min(elem["levels"])}-{max(elem["levels"])}')
        cell_level.font = normal_font
        cell_level.alignment = center_align
        cell_level.border = thin_border
        
        cell_status = ws.cell(row=row, column=5, value='必须' if is_required else '不要求')
        cell_status.font = normal_font
        cell_status.alignment = center_align
        cell_status.border = thin_border
        if is_required:
            cell_status.fill = required_fill
        
        # 验证要点
        verification_text = '\n'.join([f'□ {vp}' for vp in elem['verification_points']])
        cell_verify = ws.cell(row=row, column=6, value=verification_text)
        cell_verify.font = normal_font
        cell_verify.alignment = left_align
        cell_verify.border = thin_border
        
        cell_notes = ws.cell(row=row, column=7, value='')
        cell_notes.font = normal_font
        cell_notes.alignment = left_align
        cell_notes.border = thin_border
        
        cell_owner = ws.cell(row=row, column=8, value='')
        cell_owner.font = normal_font
        cell_owner.alignment = center_align
        cell_owner.border = thin_border
        
        ws.row_dimensions[row].height = max(20, len(elem['verification_points']) * 18)
        row += 1
    
    # 冻结首行
    ws.freeze_panes = 'A5'
    
    # 保存
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)
    print(f"PPAP检查清单已保存到: {output_path}")


def generate_checklist_json(level, customer, output_path):
    """生成PPAP检查清单JSON文件"""
    checklist = {
        'title': f'PPAP检查清单 - Level {level}',
        'customer': customer,
        'date': datetime.now().strftime('%Y-%m-%d'),
        'level': level,
        'elements': []
    }
    
    for elem in PPAP_ELEMENTS:
        is_required = level in elem['levels']
        checklist['elements'].append({
            'id': elem['id'],
            'name': elem['name'],
            'name_en': elem['name_en'],
            'required': is_required,
            'description': elem['description'],
            'verification_points': elem['verification_points']
        })
    
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(checklist, f, ensure_ascii=False, indent=2)
    print(f"PPAP检查清单(JSON)已保存到: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='PPAP检查清单生成工具')
    parser.add_argument('--level', type=int, required=True, choices=[1, 2, 3, 4, 5], help='PPAP提交等级')
    parser.add_argument('--customer', default='未指定', help='客户/OEM名称')
    parser.add_argument('--output', default='./ppap_checklist.xlsx', help='输出文件路径')
    parser.add_argument('--format', choices=['xlsx', 'json', 'both'], default='xlsx', help='输出格式')
    
    args = parser.parse_args()
    
    if args.format in ('xlsx', 'both'):
        generate_checklist_xlsx(args.level, args.customer, args.output)
    
    if args.format in ('json', 'both'):
        json_path = args.output.replace('.xlsx', '.json')
        generate_checklist_json(args.level, args.customer, json_path)


if __name__ == '__main__':
    main()
